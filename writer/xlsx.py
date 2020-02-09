from writer import IWriter
from openpyxl import Workbook


class XlsxWriter(IWriter):
    suffix = ".xlsx"

    def __init__(self, filename=''):
        super().__init__(filename=filename)
        self.filename = self.filename + self.suffix

    def write(self, data):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        for row in data:
            ws1.append(row)
        wb.save(self.filename)

    def write_multi(self, data):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = data[0]['title']
        for row in data[0]['data']:
            ws1.append(row)
        for idx, row in enumerate(data):
            if idx == 0:
                continue
            ws = wb.create_sheet(index = idx, title = row['title'])
            for row in row['data']:
                ws.append(row)
        wb.save(self.filename)